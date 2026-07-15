from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "teacher-team.xlsx"
ASSET_DIR = ROOT / "assets"
AVATAR_DIR = ASSET_DIR / "avatars"


def image_bytes_by_row(sheet) -> dict[int, bytes]:
    images: dict[int, bytes] = {}
    for image in getattr(sheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        start = getattr(anchor, "_from", None)
        if start is None:
            continue
        images[start.row + 1] = image._data()
    return images


def teacher_rows(sheet) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    current_subject = ""

    for row in range(2, sheet.max_row + 1):
        subject = sheet.cell(row, 1).value
        name = sheet.cell(row, 2).value
        intro = sheet.cell(row, 3).value

        if subject:
            current_subject = str(subject).strip()
        if not name:
            continue
        if not current_subject:
            raise RuntimeError(f"Row {row} has a teacher name but no subject context.")

        rows.append(
            {
                "row": row,
                "subject": current_subject,
                "name": str(name).strip(),
                "intro": str(intro or "").strip(),
            }
        )

    return rows


def closest_image(row: int, available_rows: set[int]) -> int:
    if not available_rows:
        raise RuntimeError(f"No avatar images remain for row {row}.")
    return min(available_rows, key=lambda image_row: (abs(image_row - row), image_row))


def build() -> None:
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook["Sheet1"]
    rows = teacher_rows(sheet)
    images = image_bytes_by_row(sheet)
    available_image_rows = set(images)

    if AVATAR_DIR.exists():
        shutil.rmtree(AVATAR_DIR)
    AVATAR_DIR.mkdir(parents=True, exist_ok=True)

    subjects: list[str] = []
    teachers: list[dict[str, str]] = []

    for item in rows:
        row = int(item["row"])
        subject = str(item["subject"])
        if subject not in subjects:
            subjects.append(subject)

        image_row = closest_image(row, available_image_rows)
        available_image_rows.remove(image_row)
        avatar_name = f"t{row}.png"
        (AVATAR_DIR / avatar_name).write_bytes(images[image_row])

        teachers.append(
            {
                "id": f"t{row}",
                "subject": subject,
                "name": str(item["name"]),
                "intro": str(item["intro"]),
                "avatar": f"assets/avatars/{avatar_name}",
            }
        )

    (ASSET_DIR / "teachers.json").write_text(
        json.dumps(
            {
                "subjects": subjects,
                "teachers": teachers,
                "defaultTitle": "新高三英才班\n剩余名额",
                "defaultNote": "注：据不完全统计，具体名额与学业诊断师沟通为准",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Generated {len(teachers)} teachers and {len(subjects)} subjects.")


if __name__ == "__main__":
    build()
